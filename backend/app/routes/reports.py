"""
Reports and Statistics routes
UC08: Báo cáo, thống kê
"""

from flask import Blueprint, request, send_file
from flask_jwt_extended import jwt_required
from app.models import (
    SanPham, LoSP, KhoHang, PhieuNhapKho, PhieuXuatKho, 
    PhieuChuyenKho, HoaDon, HoaDonSP, PhieuKiemKho, DatHang
)
from app import db
from app.utils.auth import role_required
from app.utils.helpers import success_response, error_response
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_, text, desc
import io
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.units import cm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

reports_bp = Blueprint('reports', __name__)


# =============================================
# HELPER FUNCTIONS FOR EXPORT
# =============================================

def create_excel_inventory_report(data):
    """Create Excel file for inventory report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo tồn kho"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = 'BÁO CÁO TỒN KHO'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = header_alignment
    
    # Date
    ws.merge_cells('A2:I2')
    ws['A2'] = f'Ngày tạo: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].alignment = header_alignment
    
    # Headers
    headers = ['Mã kho', 'Mã SP', 'Tên sản phẩm', 'Loại', 'ĐVT', 'Số lô', 'Tồn kho', 'HSD gần nhất', 'Trạng thái']
    ws.append([])  # Empty row
    ws.append(headers)
    
    # Style headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for item in data['inventory']:
        ws.append([
            item['MaKho'],
            item['MaSP'],
            item['TenSP'],
            item['LoaiSP'],
            item['DVT'],
            item['total_batches'],
            item['total_stock'],
            item['earliest_expiry'] or '',
            item['status']
        ])
    
    # Summary
    ws.append([])
    ws.append(['TỔNG KẾT'])
    ws.append(['Tổng sản phẩm:', data['summary']['total_products']])
    ws.append(['Tổng tồn kho:', data['summary']['total_stock']])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to bytes
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def create_excel_sales_report(data):
    """Create Excel file for sales report"""
    wb = Workbook()
    
    # Sheet 1: Summary
    ws1 = wb.active
    ws1.title = "Tổng quan"
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws1['A1'] = 'BÁO CÁO BÁN HÀNG'
    ws1['A1'].font = Font(bold=True, size=16)
    ws1.merge_cells('A1:D1')
    
    ws1['A2'] = f'Từ ngày: {data["period"]["from_date"]} đến {data["period"]["to_date"]}'
    ws1.merge_cells('A2:D2')
    
    ws1.append([])
    ws1.append(['Chỉ tiêu', 'Giá trị'])
    ws1['A4'].fill = header_fill
    ws1['A4'].font = header_font
    ws1['B4'].fill = header_fill
    ws1['B4'].font = header_font
    
    summary = data['summary']
    ws1.append(['Tổng doanh thu', f"{summary['total_revenue']:,.0f} VNĐ"])
    ws1.append(['Tổng số lượng', summary['total_quantity']])
    ws1.append(['Số hóa đơn', summary['total_invoices']])
    ws1.append(['TB doanh thu/ngày', f"{summary['average_revenue_per_day']:,.0f} VNĐ"])
    ws1.append(['TB giá trị hóa đơn', f"{summary['average_invoice_value']:,.0f} VNĐ"])
    
    # Sheet 2: Daily Sales
    if data.get('daily_sales'):
        ws2 = wb.create_sheet(title="Theo ngày")
        ws2.append(['Ngày', 'Doanh thu', 'Số lượng', 'Số hóa đơn'])
        
        for col_num in range(1, 5):
            cell = ws2.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
        
        for sale in data['daily_sales']:
            ws2.append([
                sale['date'],
                sale['total_revenue'],
                sale['total_quantity'],
                sale['total_invoices']
            ])
    
    # Sheet 3: Top Products
    if data.get('top_products'):
        ws3 = wb.create_sheet(title="Sản phẩm bán chạy")
        ws3.append(['#', 'Mã SP', 'Tên sản phẩm', 'Giá bán', 'Đã bán', 'Doanh thu'])
        
        for col_num in range(1, 7):
            cell = ws3.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
        
        for idx, product in enumerate(data['top_products'], 1):
            ws3.append([
                idx,
                product['MaSP'],
                product['TenSP'],
                product['GiaBan'],
                product['total_quantity'],
                product['total_revenue']
            ])
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def create_excel_expiry_report(data):
    """Create Excel file for expiry report"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo HSD"
    
    header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    ws['A1'] = 'BÁO CÁO HẠN SỬ DỤNG'
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    
    ws['A2'] = f'Ngày tạo: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws.merge_cells('A2:H2')
    
    # Summary
    ws.append([])
    ws.append(['Đã hết hạn:', data['summary']['total_expired'], 'Sắp hết hạn:', data['summary']['total_expiring']])
    
    # Expired items
    if data['expired']:
        ws.append([])
        ws.append(['ĐÃ HẾT HẠN'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="C00000", size=14)
        
        headers = ['Mã kho', 'Sản phẩm', 'Mã lô', 'Barcode', 'HSD', 'SL tồn', 'Quá hạn (ngày)']
        ws.append(headers)
        
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
        
        for item in data['expired']:
            ws.append([
                item['MaKho'],
                item['TenSP'],
                item['MaLo'],
                item['MaVach'],
                item['HSD'],
                item['SLTon'],
                abs(item['days_to_expiry'])
            ])
    
    # Expiring soon
    if data['expiring']:
        ws.append([])
        ws.append(['SẮP HẾT HẠN'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True, color="FF6600", size=14)
        
        headers = ['Mã kho', 'Sản phẩm', 'Mã lô', 'Barcode', 'HSD', 'SL tồn', 'Còn lại (ngày)']
        ws.append(headers)
        
        header_fill2 = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = header_fill2
            cell.font = header_font
        
        for item in data['expiring']:
            ws.append([
                item['MaKho'],
                item['TenSP'],
                item['MaLo'],
                item['MaVach'],
                item['HSD'],
                item['SLTon'],
                item['days_to_expiry']
            ])
    
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def create_pdf_simple_report(title, data, headers, filename):
    """Create a simple PDF report"""
    buffer = io.BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    
    # Title
    p.setFont("Helvetica-Bold", 16)
    p.drawString(100, height - 50, title)
    
    # Date
    p.setFont("Helvetica", 10)
    p.drawString(100, height - 70, f'Ngay tao: {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    
    # Note about Vietnamese characters
    p.setFont("Helvetica", 8)
    p.drawString(100, height - 85, '(Xuat Excel de xem day du tieng Viet)')
    
    # Simple data display
    y_position = height - 120
    p.setFont("Helvetica", 10)
    p.drawString(100, y_position, f'Total records: {len(data)}')
    
    p.showPage()
    p.save()
    
    buffer.seek(0)
    return buffer


reports_bp = Blueprint('reports', __name__)


# =============================================
# UC08: BÁO CÁO TỒN KHO
# =============================================

@reports_bp.route('/inventory', methods=['GET'])
@jwt_required()
def get_inventory_report():
    """
    Báo cáo tồn kho theo kho, theo sản phẩm, theo lô
    
    Query params:
        - ma_kho: Filter by warehouse (optional)
        - ma_sp: Filter by product (optional)
        - from_date: From date (optional)
        - to_date: To date (optional)
    """
    try:
        ma_kho = request.args.get('ma_kho')
        ma_sp = request.args.get('ma_sp')
        from_date = request.args.get('from_date')
        to_date = request.args.get('to_date')
        
        # Build query
        query = db.session.query(
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP,
            SanPham.LoaiSP,
            SanPham.DVT,
            func.count(LoSP.MaLo).label('total_batches'),
            func.sum(LoSP.SLTon).label('total_stock'),
            func.min(LoSP.HSD).label('earliest_expiry')
        ).join(SanPham, LoSP.MaSP == SanPham.MaSP)
        
        # Apply filters
        if ma_kho:
            query = query.filter(LoSP.MaKho == ma_kho)
        if ma_sp:
            query = query.filter(LoSP.MaSP == ma_sp)
        
        # Group by
        query = query.group_by(
            LoSP.MaKho, 
            LoSP.MaSP, 
            SanPham.TenSP, 
            SanPham.LoaiSP, 
            SanPham.DVT
        )
        
        results = query.all()
        
        # Format results
        inventory_data = []
        total_stock = 0
        total_products = 0
        
        for row in results:
            days_to_expiry = None
            if row.earliest_expiry:
                days_to_expiry = (row.earliest_expiry - datetime.utcnow().date()).days
            
            inventory_data.append({
                'MaKho': row.MaKho,
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'LoaiSP': row.LoaiSP,
                'DVT': row.DVT,
                'total_batches': row.total_batches,
                'total_stock': row.total_stock,
                'earliest_expiry': row.earliest_expiry.isoformat() if row.earliest_expiry else None,
                'days_to_expiry': days_to_expiry,
                'status': 'critical' if days_to_expiry and days_to_expiry <= 7 else 'warning' if days_to_expiry and days_to_expiry <= 30 else 'normal'
            })
            
            total_stock += row.total_stock
            total_products += 1
        
        return success_response({
            'inventory': inventory_data,
            'summary': {
                'total_products': total_products,
                'total_stock': total_stock,
                'total_items': len(inventory_data)
            }
        })
        
    except Exception as e:
        print(f"Inventory report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating inventory report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO XUẤT NHẬP TỒN
# =============================================

@reports_bp.route('/warehouse-movements', methods=['GET'])
@jwt_required()
def get_warehouse_movements():
    """
    Báo cáo xuất nhập tồn theo thời gian
    
    Query params:
        - from_date: From date (required)
        - to_date: To date (required)
        - ma_kho: Filter by warehouse (optional)
    """
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        ma_kho = request.args.get('ma_kho')
        
        if not from_date_str or not to_date_str:
            return error_response("from_date and to_date are required", 400)
        
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        # Get imports
        imports_query = db.session.query(
            PhieuNhapKho.MaPhieu,
            PhieuNhapKho.NgayTao,
            PhieuNhapKho.MucDich,
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP,
            func.sum(LoSP.SLTon).label('SoLuong')
        ).join(LoSP, PhieuNhapKho.MaPhieu == LoSP.MaPhieuNK)\
         .join(SanPham, LoSP.MaSP == SanPham.MaSP)\
         .filter(and_(
             PhieuNhapKho.NgayTao >= from_date,
             PhieuNhapKho.NgayTao < to_date
         ))
        
        if ma_kho:
            imports_query = imports_query.filter(LoSP.MaKho == ma_kho)
        
        imports_query = imports_query.group_by(
            PhieuNhapKho.MaPhieu,
            PhieuNhapKho.NgayTao,
            PhieuNhapKho.MucDich,
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP
        )
        
        imports = imports_query.all()
        
        # Get exports
        exports_query = db.session.query(
            PhieuXuatKho.MaPhieu,
            PhieuXuatKho.NgayTao,
            PhieuXuatKho.MucDich,
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP,
            func.sum(LoSP.SLTon).label('SoLuong')
        ).join(LoSP, PhieuXuatKho.MaPhieu == LoSP.MaPhieuXK)\
         .join(SanPham, LoSP.MaSP == SanPham.MaSP)\
         .filter(and_(
             PhieuXuatKho.NgayTao >= from_date,
             PhieuXuatKho.NgayTao < to_date
         ))
        
        if ma_kho:
            exports_query = exports_query.filter(LoSP.MaKho == ma_kho)
        
        exports_query = exports_query.group_by(
            PhieuXuatKho.MaPhieu,
            PhieuXuatKho.NgayTao,
            PhieuXuatKho.MucDich,
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP
        )
        
        exports = exports_query.all()
        
        # Format data
        movements = {
            'imports': [{
                'MaPhieu': imp.MaPhieu,
                'NgayTao': imp.NgayTao.isoformat(),
                'MucDich': imp.MucDich,
                'MaKho': imp.MaKho,
                'MaSP': imp.MaSP,
                'TenSP': imp.TenSP,
                'SoLuong': imp.SoLuong,
                'type': 'import'
            } for imp in imports],
            'exports': [{
                'MaPhieu': exp.MaPhieu,
                'NgayTao': exp.NgayTao.isoformat(),
                'MucDich': exp.MucDich,
                'MaKho': exp.MaKho,
                'MaSP': exp.MaSP,
                'TenSP': exp.TenSP,
                'SoLuong': exp.SoLuong,
                'type': 'export'
            } for exp in exports]
        }
        
        # Calculate summary
        total_imported = sum(imp.SoLuong for imp in imports)
        total_exported = sum(exp.SoLuong for exp in exports)
        
        return success_response({
            'movements': movements,
            'summary': {
                'total_imports': len(imports),
                'total_exports': len(exports),
                'total_imported_quantity': total_imported,
                'total_exported_quantity': total_exported,
                'net_change': total_imported - total_exported
            },
            'period': {
                'from_date': from_date_str,
                'to_date': to_date_str
            }
        })
        
    except Exception as e:
        print(f"Warehouse movements report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating movements report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO HẾT HẠN
# =============================================

@reports_bp.route('/expiry', methods=['GET'])
@jwt_required()
def get_expiry_report():
    """
    Báo cáo sản phẩm sắp hết hạn hoặc đã hết hạn
    
    Query params:
        - days: Number of days to check (default: 30)
        - ma_kho: Filter by warehouse (optional)
        - status: 'expired', 'expiring', 'all' (default: 'all')
    """
    try:
        days = request.args.get('days', 30, type=int)
        ma_kho = request.args.get('ma_kho')
        status_filter = request.args.get('status', 'all')
        
        today = datetime.utcnow().date()
        expiry_date = today + timedelta(days=days)
        
        # Build query
        query = db.session.query(
            LoSP.MaKho,
            LoSP.MaSP,
            LoSP.MaLo,
            LoSP.MaVach,
            LoSP.NSX,
            LoSP.HSD,
            LoSP.SLTon,
            SanPham.TenSP,
            SanPham.LoaiSP,
            SanPham.DVT
        ).join(SanPham, LoSP.MaSP == SanPham.MaSP)\
         .filter(LoSP.HSD.isnot(None))\
         .filter(LoSP.SLTon > 0)
        
        if ma_kho:
            query = query.filter(LoSP.MaKho == ma_kho)
        
        # Filter by status
        if status_filter == 'expired':
            query = query.filter(LoSP.HSD < today)
        elif status_filter == 'expiring':
            query = query.filter(and_(
                LoSP.HSD >= today,
                LoSP.HSD <= expiry_date
            ))
        else:  # all
            query = query.filter(LoSP.HSD <= expiry_date)
        
        query = query.order_by(LoSP.HSD.asc())
        
        results = query.all()
        
        # Format results
        expired = []
        expiring = []
        
        for row in results:
            days_to_expiry = (row.HSD - today).days
            is_expired = days_to_expiry < 0
            
            item = {
                'MaKho': row.MaKho,
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'LoaiSP': row.LoaiSP,
                'DVT': row.DVT,
                'MaLo': row.MaLo,
                'MaVach': row.MaVach,
                'NSX': row.NSX.isoformat() if row.NSX else None,
                'HSD': row.HSD.isoformat() if row.HSD else None,
                'SLTon': row.SLTon,
                'days_to_expiry': days_to_expiry,
                'is_expired': is_expired,
                'status': 'expired' if is_expired else 'critical' if days_to_expiry <= 7 else 'warning'
            }
            
            if is_expired:
                expired.append(item)
            else:
                expiring.append(item)
        
        return success_response({
            'expired': expired,
            'expiring': expiring,
            'summary': {
                'total_expired': len(expired),
                'total_expiring': len(expiring),
                'total_expired_quantity': sum(item['SLTon'] for item in expired),
                'total_expiring_quantity': sum(item['SLTon'] for item in expiring),
                'check_period_days': days
            }
        })
        
    except Exception as e:
        print(f"Expiry report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating expiry report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO BÁN HÀNG
# =============================================

@reports_bp.route('/sales', methods=['GET'])
@jwt_required()
def get_sales_report():
    """
    Báo cáo doanh thu bán hàng theo thời gian
    
    Query params:
        - from_date: From date (required)
        - to_date: To date (required)
        - group_by: 'day', 'week', 'month' (default: 'day')
    """
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        group_by = request.args.get('group_by', 'day')
        
        if not from_date_str or not to_date_str:
            return error_response("from_date and to_date are required", 400)
        
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        # Get sales data
        query = db.session.query(
            HoaDon.MaHD,
            HoaDon.NgayTao,
            HoaDonSP.MaSP,
            SanPham.TenSP,
            SanPham.GiaBan,
            HoaDonSP.SoLuong,
            (SanPham.GiaBan * HoaDonSP.SoLuong).label('ThanhTien')
        ).join(HoaDonSP, HoaDon.MaHD == HoaDonSP.MaHD)\
         .join(SanPham, HoaDonSP.MaSP == SanPham.MaSP)\
         .filter(and_(
             HoaDon.NgayTao >= from_date,
             HoaDon.NgayTao < to_date
         ))
        
        results = query.all()
        
        # Group data
        sales_by_date = {}
        product_sales = {}
        
        for row in results:
            date_key = row.NgayTao.date().isoformat()
            
            # Group by date
            if date_key not in sales_by_date:
                sales_by_date[date_key] = {
                    'date': date_key,
                    'total_revenue': 0,
                    'total_quantity': 0,
                    'total_invoices': set()
                }
            
            sales_by_date[date_key]['total_revenue'] += float(row.ThanhTien)
            sales_by_date[date_key]['total_quantity'] += row.SoLuong
            sales_by_date[date_key]['total_invoices'].add(row.MaHD)
            
            # Group by product
            if row.MaSP not in product_sales:
                product_sales[row.MaSP] = {
                    'MaSP': row.MaSP,
                    'TenSP': row.TenSP,
                    'GiaBan': float(row.GiaBan),
                    'total_quantity': 0,
                    'total_revenue': 0
                }
            
            product_sales[row.MaSP]['total_quantity'] += row.SoLuong
            product_sales[row.MaSP]['total_revenue'] += float(row.ThanhTien)
        
        # Format daily sales
        daily_sales = []
        for date_key, data in sorted(sales_by_date.items()):
            daily_sales.append({
                'date': data['date'],
                'total_revenue': data['total_revenue'],
                'total_quantity': data['total_quantity'],
                'total_invoices': len(data['total_invoices'])
            })
        
        # Sort products by revenue
        top_products = sorted(
            product_sales.values(),
            key=lambda x: x['total_revenue'],
            reverse=True
        )
        
        # Calculate summary
        total_revenue = sum(item['total_revenue'] for item in daily_sales)
        total_quantity = sum(item['total_quantity'] for item in daily_sales)
        total_invoices = len(set(row.MaHD for row in results))
        
        return success_response({
            'daily_sales': daily_sales,
            'top_products': top_products[:10],  # Top 10
            'summary': {
                'total_revenue': total_revenue,
                'total_quantity': total_quantity,
                'total_invoices': total_invoices,
                'average_revenue_per_day': total_revenue / len(daily_sales) if daily_sales else 0,
                'average_invoice_value': total_revenue / total_invoices if total_invoices > 0 else 0
            },
            'period': {
                'from_date': from_date_str,
                'to_date': to_date_str
            }
        })
        
    except Exception as e:
        print(f"Sales report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating sales report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO LỊCH SỬ LÔ HÀNG
# =============================================

@reports_bp.route('/batch-history', methods=['GET'])
@jwt_required()
def get_batch_history():
    """
    Lịch sử di chuyển của lô hàng
    
    Query params:
        - ma_lo: Batch code (required)
        - ma_sp: Product code (required)
    """
    try:
        ma_lo = request.args.get('ma_lo')
        ma_sp = request.args.get('ma_sp')
        
        if not ma_lo or not ma_sp:
            return error_response("ma_lo and ma_sp are required", 400)
        
        # Get batch info
        batch = LoSP.query.filter_by(MaSP=ma_sp, MaLo=ma_lo).first()
        
        if not batch:
            return error_response("Batch not found", 404)
        
        san_pham = SanPham.query.get(ma_sp)
        
        history = []
        
        # Import history
        if batch.MaPhieuNK:
            phieu_nhap = PhieuNhapKho.query.get(batch.MaPhieuNK)
            if phieu_nhap:
                history.append({
                    'type': 'import',
                    'date': phieu_nhap.NgayTao.isoformat(),
                    'ma_phieu': phieu_nhap.MaPhieu,
                    'muc_dich': phieu_nhap.MucDich,
                    'ma_kho': batch.MaKho,
                    'so_luong': batch.SLTon,
                    'action': 'Nhập kho'
                })
        
        # Export history
        if batch.MaPhieuXK:
            phieu_xuat = PhieuXuatKho.query.get(batch.MaPhieuXK)
            if phieu_xuat:
                history.append({
                    'type': 'export',
                    'date': phieu_xuat.NgayTao.isoformat(),
                    'ma_phieu': phieu_xuat.MaPhieu,
                    'muc_dich': phieu_xuat.MucDich,
                    'ma_kho': batch.MaKho,
                    'so_luong': batch.SLTon,
                    'action': 'Xuất kho'
                })
        
        # Sort by date
        history.sort(key=lambda x: x['date'])
        
        return success_response({
            'batch_info': {
                **batch.to_dict(),
                'product': san_pham.to_dict() if san_pham else None
            },
            'history': history,
            'total_movements': len(history)
        })
        
    except Exception as e:
        print(f"Batch history error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error getting batch history: {str(e)}", 500)


# =============================================
# UC08: DASHBOARD STATISTICS
# =============================================

@reports_bp.route('/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard_statistics():
    """Get overall statistics for dashboard"""
    try:
        # Total products
        total_products = SanPham.query.count()
        
        # Total stock
        total_stock = db.session.query(func.sum(LoSP.SLTon)).scalar() or 0
        
        # Low stock products - Fixed SQL query
        low_stock_query = text("""
            SELECT COUNT(*) as count
            FROM SanPham sp
            INNER JOIN (
                SELECT MaSP, SUM(SLTon) as total_stock
                FROM LoSP
                GROUP BY MaSP
            ) AS stock_summary ON sp.MaSP = stock_summary.MaSP
            WHERE stock_summary.total_stock < sp.MucCanhBaoDatHang
        """)
        
        low_stock_result = db.session.execute(low_stock_query).fetchone()
        low_stock_products = low_stock_result.count if low_stock_result else 0
        
        # Expired batches
        today = datetime.utcnow().date()
        expired_batches = LoSP.query.filter(
            and_(LoSP.HSD < today, LoSP.SLTon > 0)
        ).count()
        
        # Expiring soon (7 days)
        expiring_soon = LoSP.query.filter(
            and_(
                LoSP.HSD >= today,
                LoSP.HSD <= today + timedelta(days=7),
                LoSP.SLTon > 0
            )
        ).count()
        
        # Recent activities (last 7 days)
        week_ago = datetime.utcnow() - timedelta(days=7)
        
        recent_imports = PhieuNhapKho.query.filter(
            PhieuNhapKho.NgayTao >= week_ago
        ).count()
        
        recent_exports = PhieuXuatKho.query.filter(
            PhieuXuatKho.NgayTao >= week_ago
        ).count()
        
        # Build alerts
        alerts = []
        if expired_batches > 0:
            alerts.append(f"Có {expired_batches} lô hàng đã hết hạn cần xử lý")
        if expiring_soon > 0:
            alerts.append(f"Có {expiring_soon} lô hàng sắp hết hạn trong 7 ngày")
        if low_stock_products > 0:
            alerts.append(f"Có {low_stock_products} sản phẩm cần đặt hàng")
        
        return success_response({
            'total_products': total_products,
            'total_stock': total_stock,
            'low_stock': low_stock_products,
            'expiring_soon': expiring_soon,
            'recent_activities': {
                'imports_last_7_days': recent_imports,
                'exports_last_7_days': recent_exports
            },
            'alerts': alerts
        })
        
    except Exception as e:
        print(f"Dashboard statistics error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error getting dashboard statistics: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO TRẢ HÀNG
# =============================================

@reports_bp.route('/returns', methods=['GET'])
@jwt_required()
def get_returns_report():
    """
    Báo cáo trả hàng theo thời gian
    
    Query params:
        - from_date: From date (optional)
        - to_date: To date (optional)
    """
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        # Build query for returns (Phiếu nhập có mục đích trả hàng)
        query = db.session.query(
            PhieuNhapKho.MaPhieu,
            PhieuNhapKho.NgayTao,
            PhieuNhapKho.MucDich,
            PhieuNhapKho.MaThamChieu,
            LoSP.MaKho,
            LoSP.MaSP,
            LoSP.MaLo,
            SanPham.TenSP,
            SanPham.GiaBan,
            LoSP.SLTon
        ).join(LoSP, PhieuNhapKho.MaPhieu == LoSP.MaPhieuNK)\
         .join(SanPham, LoSP.MaSP == SanPham.MaSP)\
         .filter(PhieuNhapKho.MucDich.like('%trả hàng%'))
        
        # Apply date filters
        if from_date_str:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            query = query.filter(PhieuNhapKho.NgayTao >= from_date)
        
        if to_date_str:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PhieuNhapKho.NgayTao < to_date)
        
        query = query.order_by(desc(PhieuNhapKho.NgayTao))
        
        results = query.all()
        
        # Format results
        returns = []
        total_returned_quantity = 0
        total_returned_value = 0
        
        for row in results:
            returned_value = float(row.GiaBan * row.SLTon)
            returns.append({
                'MaPhieu': row.MaPhieu,
                'NgayTao': row.NgayTao.isoformat(),
                'MucDich': row.MucDich,
                'MaThamChieu': row.MaThamChieu,
                'MaKho': row.MaKho,
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'MaLo': row.MaLo,
                'SoLuong': row.SLTon,
                'GiaBan': float(row.GiaBan),
                'GiaTriTra': returned_value
            })
            
            total_returned_quantity += row.SLTon
            total_returned_value += returned_value
        
        return success_response({
            'returns': returns,
            'summary': {
                'total_returns': len(returns),
                'total_returned_quantity': total_returned_quantity,
                'total_returned_value': total_returned_value
            }
        })
        
    except Exception as e:
        print(f"Returns report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating returns report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO HOẠT ĐỘNG KHO
# =============================================

@reports_bp.route('/warehouse-activities', methods=['GET'])
@jwt_required()
def get_warehouse_activities():
    """
    Báo cáo hoạt động kho tổng hợp
    
    Query params:
        - from_date: From date (optional, default: 7 days ago)
        - to_date: To date (optional, default: now)
        - ma_kho: Filter by warehouse (optional)
    """
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        ma_kho = request.args.get('ma_kho')
        
        # Default to last 7 days
        if not from_date_str:
            from_date = datetime.utcnow() - timedelta(days=7)
        else:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        
        if not to_date_str:
            to_date = datetime.utcnow()
        else:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        activities = []
        
        # Get imports
        imports_query = PhieuNhapKho.query.filter(
            and_(
                PhieuNhapKho.NgayTao >= from_date,
                PhieuNhapKho.NgayTao < to_date
            )
        )
        
        for phieu in imports_query.all():
            # Get batches for this import
            batches = LoSP.query.filter_by(MaPhieuNK=phieu.MaPhieu)
            
            if ma_kho:
                batches = batches.filter_by(MaKho=ma_kho)
            
            batches = batches.all()
            
            if batches:
                total_quantity = sum(b.SLTon for b in batches)
                activities.append({
                    'type': 'import',
                    'MaPhieu': phieu.MaPhieu,
                    'NgayTao': phieu.NgayTao.isoformat(),
                    'MucDich': phieu.MucDich,
                    'MaThamChieu': phieu.MaThamChieu,
                    'MaKho': batches[0].MaKho if batches else None,
                    'total_quantity': total_quantity,
                    'batches_count': len(batches)
                })
        
        # Get exports
        exports_query = PhieuXuatKho.query.filter(
            and_(
                PhieuXuatKho.NgayTao >= from_date,
                PhieuXuatKho.NgayTao < to_date
            )
        )
        
        for phieu in exports_query.all():
            # Get batches for this export
            batches = LoSP.query.filter_by(MaPhieuXK=phieu.MaPhieu)
            
            if ma_kho:
                batches = batches.filter_by(MaKho=ma_kho)
            
            batches = batches.all()
            
            if batches:
                total_quantity = sum(b.SLTon for b in batches)
                activities.append({
                    'type': 'export',
                    'MaPhieu': phieu.MaPhieu,
                    'NgayTao': phieu.NgayTao.isoformat(),
                    'MucDich': phieu.MucDich,
                    'MaThamChieu': phieu.MaThamChieu,
                    'MaKho': batches[0].MaKho if batches else None,
                    'total_quantity': total_quantity,
                    'batches_count': len(batches)
                })
        
        # Get transfers
        transfers_query = PhieuChuyenKho.query.filter(
            and_(
                PhieuChuyenKho.NgayTao >= from_date,
                PhieuChuyenKho.NgayTao < to_date
            )
        )
        
        if ma_kho:
            transfers_query = transfers_query.filter(
                or_(
                    PhieuChuyenKho.KhoXuat == ma_kho,
                    PhieuChuyenKho.KhoNhap == ma_kho
                )
            )
        
        for phieu in transfers_query.all():
            activities.append({
                'type': 'transfer',
                'MaPhieu': phieu.MaPhieu,
                'NgayTao': phieu.NgayTao.isoformat(),
                'MucDich': phieu.MucDich,
                'MaThamChieu': phieu.MaThamChieu,
                'KhoXuat': phieu.KhoXuat,
                'KhoNhap': phieu.KhoNhap
            })
        
        # Get inventory checks
        inventory_query = PhieuKiemKho.query.filter(
            and_(
                PhieuKiemKho.NgayTao >= from_date,
                PhieuKiemKho.NgayTao < to_date
            )
        )
        
        if ma_kho:
            inventory_query = inventory_query.filter_by(MaKho=ma_kho)
        
        for phieu in inventory_query.all():
            activities.append({
                'type': 'inventory_check',
                'MaPhieu': phieu.MaPhieu,
                'NgayTao': phieu.NgayTao.isoformat(),
                'MucDich': phieu.MucDich,
                'MaKho': phieu.MaKho
            })
        
        # Sort by date
        activities.sort(key=lambda x: x['NgayTao'], reverse=True)
        
        # Calculate summary
        total_imports = sum(1 for a in activities if a['type'] == 'import')
        total_exports = sum(1 for a in activities if a['type'] == 'export')
        total_transfers = sum(1 for a in activities if a['type'] == 'transfer')
        total_checks = sum(1 for a in activities if a['type'] == 'inventory_check')
        
        return success_response({
            'activities': activities,
            'summary': {
                'total_activities': len(activities),
                'total_imports': total_imports,
                'total_exports': total_exports,
                'total_transfers': total_transfers,
                'total_inventory_checks': total_checks
            },
            'period': {
                'from_date': from_date.date().isoformat(),
                'to_date': to_date.date().isoformat()
            }
        })
        
    except Exception as e:
        print(f"Warehouse activities error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error getting warehouse activities: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO ĐƠN ĐẶT HÀNG
# =============================================

@reports_bp.route('/supplier-orders', methods=['GET'])
@jwt_required()
def get_supplier_orders_report():
    """
    Báo cáo đơn đặt hàng từ nhà cung cấp
    
    Query params:
        - trang_thai: Filter by status (optional)
        - from_date: From date (optional)
        - to_date: To date (optional)
    """
    try:
        trang_thai = request.args.get('trang_thai')
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        query = DatHang.query
        
        # Apply filters
        if trang_thai:
            query = query.filter_by(TrangThai=trang_thai)
        
        if from_date_str:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
            query = query.filter(DatHang.NgayDat >= from_date)
        
        if to_date_str:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(DatHang.NgayDat < to_date)
        
        query = query.order_by(desc(DatHang.NgayDat))
        
        results = query.all()
        
        # Format results
        orders = []
        total_orders = 0
        total_value = 0
        
        for order in results:
            order_data = order.to_dict()
            
            # Calculate order value from ChiTietDonHang JSON
            order_value = 0
            if order.ChiTietDonHang:
                import json
                chi_tiet = json.loads(order.ChiTietDonHang) if isinstance(order.ChiTietDonHang, str) else order.ChiTietDonHang
                for item in chi_tiet:
                    order_value += item.get('SoLuong', 0) * item.get('DonGia', 0)
            
            order_data['order_value'] = order_value
            orders.append(order_data)
            
            total_orders += 1
            total_value += order_value
        
        # Group by status
        by_status = {}
        for order in orders:
            status = order['TrangThai']
            if status not in by_status:
                by_status[status] = {'count': 0, 'value': 0}
            by_status[status]['count'] += 1
            by_status[status]['value'] += order['order_value']
        
        return success_response({
            'orders': orders,
            'summary': {
                'total_orders': total_orders,
                'total_value': total_value,
                'by_status': by_status
            }
        })
        
    except Exception as e:
        print(f"Supplier orders report error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating supplier orders report: {str(e)}", 500)


# =============================================
# UC08: BÁO CÁO SẢN PHẨM BÁN CHẠY
# =============================================

@reports_bp.route('/top-products', methods=['GET'])
@jwt_required()
def get_top_products():
    """
    Báo cáo sản phẩm bán chạy
    
    Query params:
        - from_date: From date (optional, default: 30 days ago)
        - to_date: To date (optional, default: now)
        - limit: Number of products to return (default: 10)
    """
    try:
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        limit = request.args.get('limit', 10, type=int)
        
        # Default to last 30 days
        if not from_date_str:
            from_date = datetime.utcnow() - timedelta(days=30)
        else:
            from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        
        if not to_date_str:
            to_date = datetime.utcnow()
        else:
            to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        # Query top selling products
        query = db.session.query(
            SanPham.MaSP,
            SanPham.TenSP,
            SanPham.LoaiSP,
            SanPham.DVT,
            SanPham.GiaBan,
            func.sum(HoaDonSP.SoLuong).label('total_quantity'),
            func.count(HoaDonSP.MaHD).label('total_invoices'),
            (func.sum(HoaDonSP.SoLuong) * SanPham.GiaBan).label('total_revenue')
        ).join(HoaDonSP, SanPham.MaSP == HoaDonSP.MaSP)\
         .join(HoaDon, HoaDonSP.MaHD == HoaDon.MaHD)\
         .filter(and_(
             HoaDon.NgayTao >= from_date,
             HoaDon.NgayTao < to_date
         ))\
         .group_by(
             SanPham.MaSP,
             SanPham.TenSP,
             SanPham.LoaiSP,
             SanPham.DVT,
             SanPham.GiaBan
         )\
         .order_by(desc('total_quantity'))\
         .limit(limit)
        
        results = query.all()
        
        # Format results
        top_products = []
        for row in results:
            top_products.append({
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'LoaiSP': row.LoaiSP,
                'DVT': row.DVT,
                'GiaBan': float(row.GiaBan),
                'total_quantity': row.total_quantity,
                'total_invoices': row.total_invoices,
                'total_revenue': float(row.total_revenue)
            })
        
        return success_response({
            'top_products': top_products,
            'period': {
                'from_date': from_date.date().isoformat(),
                'to_date': to_date.date().isoformat()
            }
        })
        
    except Exception as e:
        print(f"Top products error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error getting top products: {str(e)}", 500)


# =============================================
# UC08: DỰ BÁO HẾT HÀNG
# =============================================

@reports_bp.route('/stock-forecast', methods=['GET'])
@jwt_required()
def get_stock_forecast():
    """
    Dự báo thời gian hết hàng dựa trên tốc độ bán
    
    Query params:
        - ma_sp: Product code (optional)
        - days: Number of days to analyze sales velocity (default: 30)
    """
    try:
        ma_sp = request.args.get('ma_sp')
        days = request.args.get('days', 30, type=int)
        
        # Calculate date range
        to_date = datetime.utcnow()
        from_date = to_date - timedelta(days=days)
        
        # Get products query
        products_query = SanPham.query
        if ma_sp:
            products_query = products_query.filter_by(MaSP=ma_sp)
        
        products = products_query.all()
        
        forecasts = []
        
        for product in products:
            # Get current stock
            current_stock = db.session.query(
                func.sum(LoSP.SLTon)
            ).filter_by(MaSP=product.MaSP).scalar() or 0
            
            if current_stock == 0:
                continue
            
            # Get sales in the period
            total_sold = db.session.query(
                func.sum(HoaDonSP.SoLuong)
            ).join(HoaDon, HoaDonSP.MaHD == HoaDon.MaHD)\
             .filter(
                 and_(
                     HoaDonSP.MaSP == product.MaSP,
                     HoaDon.NgayTao >= from_date,
                     HoaDon.NgayTao < to_date
                 )
             ).scalar() or 0
            
            # Calculate daily average
            daily_average = total_sold / days if days > 0 else 0
            
            # Calculate days until out of stock
            if daily_average > 0:
                days_until_out = current_stock / daily_average
            else:
                days_until_out = None
            
            # Calculate suggested order quantity
            suggested_order = max(0, product.MucCanhBaoDatHang - current_stock)
            
            forecasts.append({
                'MaSP': product.MaSP,
                'TenSP': product.TenSP,
                'LoaiSP': product.LoaiSP,
                'current_stock': current_stock,
                'daily_average_sales': round(daily_average, 2),
                'days_until_out_of_stock': round(days_until_out, 1) if days_until_out else None,
                'reorder_level': product.MucCanhBaoDatHang,
                'suggested_order_quantity': suggested_order,
                'needs_reorder': current_stock < product.MucCanhBaoDatHang,
                'urgency': 'critical' if days_until_out and days_until_out <= 7 else 'warning' if days_until_out and days_until_out <= 14 else 'normal'
            })
        
        # Sort by urgency
        forecasts.sort(key=lambda x: (
            0 if x['urgency'] == 'critical' else 1 if x['urgency'] == 'warning' else 2,
            x['days_until_out_of_stock'] if x['days_until_out_of_stock'] else 999
        ))
        
        return success_response({
            'forecasts': forecasts,
            'analysis_period': {
                'from_date': from_date.date().isoformat(),
                'to_date': to_date.date().isoformat(),
                'days': days
            },
            'summary': {
                'total_products': len(forecasts),
                'critical_products': sum(1 for f in forecasts if f['urgency'] == 'critical'),
                'needs_reorder': sum(1 for f in forecasts if f['needs_reorder'])
            }
        })
        
    except Exception as e:
        print(f"Stock forecast error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error generating stock forecast: {str(e)}", 500)


# =============================================
# UC08: EXPORT REPORTS TO FILE (PDF/EXCEL)
# =============================================

@reports_bp.route('/export/inventory', methods=['GET'])
@jwt_required()
def export_inventory_report():
    """
    Export inventory report to Excel or PDF
    
    Query params:
        - format: 'excel' or 'pdf' (default: 'excel')
        - ma_kho: Filter by warehouse (optional)
        - ma_sp: Filter by product (optional)
    """
    try:
        format_type = request.args.get('format', 'excel').lower()
        
        # Get report data (reuse existing function logic)
        ma_kho = request.args.get('ma_kho')
        ma_sp = request.args.get('ma_sp')
        
        query = db.session.query(
            LoSP.MaKho,
            LoSP.MaSP,
            SanPham.TenSP,
            SanPham.LoaiSP,
            SanPham.DVT,
            func.count(LoSP.MaLo).label('total_batches'),
            func.sum(LoSP.SLTon).label('total_stock'),
            func.min(LoSP.HSD).label('earliest_expiry')
        ).join(SanPham, LoSP.MaSP == SanPham.MaSP)
        
        if ma_kho:
            query = query.filter(LoSP.MaKho == ma_kho)
        if ma_sp:
            query = query.filter(LoSP.MaSP == ma_sp)
        
        query = query.group_by(
            LoSP.MaKho, LoSP.MaSP, SanPham.TenSP, SanPham.LoaiSP, SanPham.DVT
        )
        
        results = query.all()
        
        inventory_data = []
        total_stock = 0
        total_products = 0
        
        for row in results:
            days_to_expiry = None
            if row.earliest_expiry:
                days_to_expiry = (row.earliest_expiry - datetime.utcnow().date()).days
            
            inventory_data.append({
                'MaKho': row.MaKho,
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'LoaiSP': row.LoaiSP,
                'DVT': row.DVT,
                'total_batches': row.total_batches,
                'total_stock': row.total_stock,
                'earliest_expiry': row.earliest_expiry.isoformat() if row.earliest_expiry else None,
                'days_to_expiry': days_to_expiry,
                'status': 'critical' if days_to_expiry and days_to_expiry <= 7 else 'warning' if days_to_expiry and days_to_expiry <= 30 else 'normal'
            })
            
            total_stock += row.total_stock
            total_products += 1
        
        data = {
            'inventory': inventory_data,
            'summary': {
                'total_products': total_products,
                'total_stock': total_stock,
                'total_items': len(inventory_data)
            }
        }
        
        if format_type == 'excel':
            excel_file = create_excel_inventory_report(data)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'bao_cao_ton_kho_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        elif format_type == 'pdf':
            pdf_file = create_pdf_simple_report(
                'BAO CAO TON KHO',
                inventory_data,
                ['MaKho', 'MaSP', 'TenSP'],
                'inventory'
            )
            return send_file(
                pdf_file,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'bao_cao_ton_kho_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
        else:
            return error_response("Invalid format. Use 'excel' or 'pdf'", 400)
            
    except Exception as e:
        print(f"Export inventory error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error exporting inventory report: {str(e)}", 500)


@reports_bp.route('/export/sales', methods=['GET'])
@jwt_required()
def export_sales_report():
    """
    Export sales report to Excel or PDF
    
    Query params:
        - format: 'excel' or 'pdf' (default: 'excel')
        - from_date: From date (required)
        - to_date: To date (required)
    """
    try:
        format_type = request.args.get('format', 'excel').lower()
        from_date_str = request.args.get('from_date')
        to_date_str = request.args.get('to_date')
        
        if not from_date_str or not to_date_str:
            return error_response("from_date and to_date are required", 400)
        
        from_date = datetime.strptime(from_date_str, '%Y-%m-%d')
        to_date = datetime.strptime(to_date_str, '%Y-%m-%d') + timedelta(days=1)
        
        # Get sales data
        query = db.session.query(
            HoaDon.MaHD,
            HoaDon.NgayTao,
            HoaDonSP.MaSP,
            SanPham.TenSP,
            SanPham.GiaBan,
            HoaDonSP.SoLuong,
            (SanPham.GiaBan * HoaDonSP.SoLuong).label('ThanhTien')
        ).join(HoaDonSP, HoaDon.MaHD == HoaDonSP.MaHD)\
         .join(SanPham, HoaDonSP.MaSP == SanPham.MaSP)\
         .filter(and_(
             HoaDon.NgayTao >= from_date,
             HoaDon.NgayTao < to_date
         ))
        
        results = query.all()
        
        sales_by_date = {}
        product_sales = {}
        
        for row in results:
            date_key = row.NgayTao.date().isoformat()
            
            if date_key not in sales_by_date:
                sales_by_date[date_key] = {
                    'date': date_key,
                    'total_revenue': 0,
                    'total_quantity': 0,
                    'total_invoices': set()
                }
            
            sales_by_date[date_key]['total_revenue'] += float(row.ThanhTien)
            sales_by_date[date_key]['total_quantity'] += row.SoLuong
            sales_by_date[date_key]['total_invoices'].add(row.MaHD)
            
            if row.MaSP not in product_sales:
                product_sales[row.MaSP] = {
                    'MaSP': row.MaSP,
                    'TenSP': row.TenSP,
                    'GiaBan': float(row.GiaBan),
                    'total_quantity': 0,
                    'total_revenue': 0
                }
            
            product_sales[row.MaSP]['total_quantity'] += row.SoLuong
            product_sales[row.MaSP]['total_revenue'] += float(row.ThanhTien)
        
        daily_sales = []
        for date_key, data in sorted(sales_by_date.items()):
            daily_sales.append({
                'date': data['date'],
                'total_revenue': data['total_revenue'],
                'total_quantity': data['total_quantity'],
                'total_invoices': len(data['total_invoices'])
            })
        
        top_products = sorted(
            product_sales.values(),
            key=lambda x: x['total_revenue'],
            reverse=True
        )[:10]
        
        total_revenue = sum(item['total_revenue'] for item in daily_sales)
        total_quantity = sum(item['total_quantity'] for item in daily_sales)
        total_invoices = len(set(row.MaHD for row in results))
        
        data = {
            'daily_sales': daily_sales,
            'top_products': top_products,
            'summary': {
                'total_revenue': total_revenue,
                'total_quantity': total_quantity,
                'total_invoices': total_invoices,
                'average_revenue_per_day': total_revenue / len(daily_sales) if daily_sales else 0,
                'average_invoice_value': total_revenue / total_invoices if total_invoices > 0 else 0
            },
            'period': {
                'from_date': from_date_str,
                'to_date': to_date_str
            }
        }
        
        if format_type == 'excel':
            excel_file = create_excel_sales_report(data)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'bao_cao_ban_hang_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        elif format_type == 'pdf':
            pdf_file = create_pdf_simple_report(
                'BAO CAO BAN HANG',
                daily_sales,
                ['date', 'total_revenue'],
                'sales'
            )
            return send_file(
                pdf_file,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'bao_cao_ban_hang_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
        else:
            return error_response("Invalid format. Use 'excel' or 'pdf'", 400)
            
    except Exception as e:
        print(f"Export sales error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error exporting sales report: {str(e)}", 500)


@reports_bp.route('/export/expiry', methods=['GET'])
@jwt_required()
def export_expiry_report():
    """
    Export expiry report to Excel or PDF
    
    Query params:
        - format: 'excel' or 'pdf' (default: 'excel')
        - days: Number of days to check (default: 30)
        - ma_kho: Filter by warehouse (optional)
        - status: 'expired', 'expiring', 'all' (default: 'all')
    """
    try:
        format_type = request.args.get('format', 'excel').lower()
        days = request.args.get('days', 30, type=int)
        ma_kho = request.args.get('ma_kho')
        status_filter = request.args.get('status', 'all')
        
        today = datetime.utcnow().date()
        expiry_date = today + timedelta(days=days)
        
        query = db.session.query(
            LoSP.MaKho,
            LoSP.MaSP,
            LoSP.MaLo,
            LoSP.MaVach,
            LoSP.NSX,
            LoSP.HSD,
            LoSP.SLTon,
            SanPham.TenSP,
            SanPham.LoaiSP,
            SanPham.DVT
        ).join(SanPham, LoSP.MaSP == SanPham.MaSP)\
         .filter(LoSP.HSD.isnot(None))\
         .filter(LoSP.SLTon > 0)
        
        if ma_kho:
            query = query.filter(LoSP.MaKho == ma_kho)
        
        if status_filter == 'expired':
            query = query.filter(LoSP.HSD < today)
        elif status_filter == 'expiring':
            query = query.filter(and_(
                LoSP.HSD >= today,
                LoSP.HSD <= expiry_date
            ))
        else:
            query = query.filter(LoSP.HSD <= expiry_date)
        
        query = query.order_by(LoSP.HSD.asc())
        results = query.all()
        
        expired = []
        expiring = []
        
        for row in results:
            days_to_expiry = (row.HSD - today).days
            is_expired = days_to_expiry < 0
            
            item = {
                'MaKho': row.MaKho,
                'MaSP': row.MaSP,
                'TenSP': row.TenSP,
                'LoaiSP': row.LoaiSP,
                'DVT': row.DVT,
                'MaLo': row.MaLo,
                'MaVach': row.MaVach,
                'NSX': row.NSX.isoformat() if row.NSX else None,
                'HSD': row.HSD.isoformat() if row.HSD else None,
                'SLTon': row.SLTon,
                'days_to_expiry': days_to_expiry,
                'is_expired': is_expired,
                'status': 'expired' if is_expired else 'critical' if days_to_expiry <= 7 else 'warning'
            }
            
            if is_expired:
                expired.append(item)
            else:
                expiring.append(item)
        
        data = {
            'expired': expired,
            'expiring': expiring,
            'summary': {
                'total_expired': len(expired),
                'total_expiring': len(expiring),
                'total_expired_quantity': sum(item['SLTon'] for item in expired),
                'total_expiring_quantity': sum(item['SLTon'] for item in expiring),
                'check_period_days': days
            }
        }
        
        if format_type == 'excel':
            excel_file = create_excel_expiry_report(data)
            return send_file(
                excel_file,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=f'bao_cao_han_su_dung_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            )
        elif format_type == 'pdf':
            all_items = expired + expiring
            pdf_file = create_pdf_simple_report(
                'BAO CAO HAN SU DUNG',
                all_items,
                ['MaKho', 'TenSP', 'HSD'],
                'expiry'
            )
            return send_file(
                pdf_file,
                mimetype='application/pdf',
                as_attachment=True,
                download_name=f'bao_cao_han_su_dung_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
            )
        else:
            return error_response("Invalid format. Use 'excel' or 'pdf'", 400)
            
    except Exception as e:
        print(f"Export expiry error: {str(e)}")
        import traceback
        traceback.print_exc()
        return error_response(f"Error exporting expiry report: {str(e)}", 500)